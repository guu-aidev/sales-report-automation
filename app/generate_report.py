"""
月別売上集計レポート自動生成スクリプト

使い方:
    python generate_report.py
    python generate_report.py --input data/sample --output output/report.xlsx --year 2024
"""
import argparse
import sqlite3
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.packaging.extended import ExtendedProperties

# openpyxl 3.1.4+ のバグ回避:
# app.xml に "Compatible / Openpyxl" が入ると Excel がチャートを誤レンダリングする
_orig_ep_init = ExtendedProperties.__init__
def _fixed_ep_init(self, *args, **kwargs):
    _orig_ep_init(self, *args, **kwargs)
    self.Application = "Microsoft Excel"
ExtendedProperties.__init__ = _fixed_ep_init
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ai_commentary import generate_kpi_commentary

MONTH_NAMES = ["1月", "2月", "3月", "4月", "5月", "6月",
               "7月", "8月", "9月", "10月", "11月", "12月"]

HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_ROW_FILL   = PatternFill("solid", fgColor="DEEAF1")
TOTAL_FILL     = PatternFill("solid", fgColor="BDD7EE")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1F4E79")
BOLD_FONT   = Font(bold=True)
THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 前月比の条件付き書式カラー（Excel標準配色）
CF_RED_FILL    = PatternFill(bgColor="FFC7CE")
CF_RED_FONT    = Font(color="9C0006", bold=True)
CF_YELLOW_FILL = PatternFill(bgColor="E2E8F0")
CF_YELLOW_FONT = Font(color="475569", bold=True)
CF_GREEN_FILL  = PatternFill(bgColor="C6EFCE")
CF_GREEN_FONT  = Font(color="276221", bold=True)

FMT_YEN   = '#,##0"円"'
FMT_COUNT = '#,##0"人"'


def _yen_fmt(value: int) -> str:
    """100万円以上は千円単位、未満は円単位のExcelフォーマット文字列を返す"""
    return '#,##0,"千円"' if value >= 1_000_000 else '#,##0"円"'


def _yen_unit(value: int) -> str:
    return "千円" if value >= 1_000_000 else "円"


def _yen_str(value: int) -> str:
    """テキスト用: 100万円以上は千円表示、未満は円表示"""
    return f"{value // 1000:,}千円" if value >= 1_000_000 else f"{value:,}円"


# ── ヘルパー ──────────────────────────────────────────────────────────────

def _read_csv_auto(path: Path) -> pd.DataFrame:
    """UTF-8 → Shift_JIS(cp932) の順に試して読み込む。
    Excelの「CSV UTF-8」も「CSV (カンマ区切り)」も両方扱えるようにする。
    """
    for enc in ("utf-8-sig", "cp932"):
        try:
            return pd.read_csv(path, encoding=enc, parse_dates=["日付"])
        except UnicodeDecodeError:
            continue
    raise ValueError(f"{path.name}: 文字コードを判定できません（UTF-8 / Shift_JIS のいずれかで保存してください）")


def load_csv_files(input_dir: Path, year: int) -> pd.DataFrame:
    frames = []
    for csv_path in sorted(input_dir.glob("*.csv")):
        df = _read_csv_auto(csv_path)
        frames.append(df)
    if not frames:
        raise FileNotFoundError(f"CSVファイルが見つかりません: {input_dir}")
    data = pd.concat(frames, ignore_index=True)
    data = data[data["日付"].dt.year == year].copy()
    data["月"] = data["日付"].dt.month
    return data


# ── DB層 ──────────────────────────────────────────────────────────────────

_BASE   = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent.parent
DB_PATH = _BASE / "data" / "sales.db"


def import_csv_to_db(input_dir: Path, db_path: Path = DB_PATH) -> None:
    """CSVをSQLiteにインポートする。
    CSVにある（日付×店舗名）の行は上書き、DBにあってCSVにない行は保持する。
    """
    REQUIRED = {"日付", "店舗名", "売上金額", "客数"}
    frames = []
    for csv_path in sorted(input_dir.glob("*.csv")):
        df = _read_csv_auto(csv_path)
        missing = REQUIRED - set(df.columns)
        if missing:
            raise ValueError(f"{csv_path.name}: 必須列が不足しています → {', '.join(sorted(missing))}")
        frames.append(df)
    if not frames:
        raise FileNotFoundError(f"CSVファイルが見つかりません: {input_dir}")
    new_data = pd.concat(frames, ignore_index=True)
    new_data["年"] = new_data["日付"].dt.year
    new_data["月"] = new_data["日付"].dt.month
    new_data["日付"] = new_data["日付"].dt.strftime("%Y-%m-%d")

    conn = sqlite3.connect(db_path)
    table_exists = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='sales'"
    ).fetchone() is not None

    if table_exists:
        existing = pd.read_sql_query("SELECT * FROM sales", conn)
        # CSVにある（日付, 店舗名）ペアだけDBから除き、残りは保持
        keys = new_data[["日付", "店舗名"]].drop_duplicates()
        merged = existing.merge(keys, on=["日付", "店舗名"], how="left", indicator=True)
        surviving = existing[merged["_merge"] == "left_only"]
        combined = pd.concat([surviving, new_data], ignore_index=True)
    else:
        combined = new_data

    combined.to_sql("sales", conn, if_exists="replace", index=False)
    conn.commit()
    conn.close()


def load_from_db(db_path: Path, year: int) -> pd.DataFrame:
    """指定年のデータをDBから取得"""
    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM sales WHERE 年 = ?", conn, params=(year,))
    conn.close()
    df["月"] = df["月"].astype(int)
    return df


def get_years_from_db(db_path: Path) -> list:
    """DBに存在する年度一覧を返す"""
    conn = sqlite3.connect(db_path)
    rows = conn.execute("SELECT DISTINCT 年 FROM sales ORDER BY 年").fetchall()
    conn.close()
    return [r[0] for r in rows]


def get_stores_from_db(db_path: Path, year: int) -> list:
    """指定年に存在する店舗一覧を返す"""
    conn = sqlite3.connect(db_path)
    rows = conn.execute(
        "SELECT DISTINCT 店舗名 FROM sales WHERE 年 = ? ORDER BY 店舗名", (year,)
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]


def build_monthly_pivot(data: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    sales_pivot = (
        data.groupby(["月", "店舗名"])["売上金額"]
        .sum().unstack(fill_value=0).reindex(range(1, 13), fill_value=0)
    )
    count_pivot = (
        data.groupby(["月", "店舗名"])["客数"]
        .sum().unstack(fill_value=0).reindex(range(1, 13), fill_value=0)
    )
    return sales_pivot, count_pivot


def style_header_row(ws, row: int, col_start: int, col_end: int, fill=None):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_data_cell(cell, is_alt: bool, number_format: str = "#,##0"):
    if is_alt:
        cell.fill = ALT_ROW_FILL
    cell.border = BORDER
    cell.number_format = number_format
    cell.alignment = Alignment(horizontal="right")


def _apply_mom_cf(ws, cell_range: str):
    """前月比セル範囲に赤/黄/緑の条件付き書式を適用"""
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="lessThan", formula=["-0.05"],
                   fill=CF_RED_FILL, font=CF_RED_FONT))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="between", formula=["-0.05", "0.05"],
                   fill=CF_YELLOW_FILL, font=CF_YELLOW_FONT))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="greaterThan", formula=["0.05"],
                   fill=CF_GREEN_FILL, font=CF_GREEN_FONT))


def _kpi_card(ws, row: int, col_start: int, col_end: int,
              label: str, value, sublabel: str, bg_hex: str,
              number_format: str = None):
    """3行構成のKPIカード: [ラベル行] [値行（大字）] [補足行]"""
    bg_fill    = PatternFill("solid", fgColor=bg_hex)
    light_fill = PatternFill("solid", fgColor="EBF3FB")
    med = Side(style="medium")
    no  = Side(style=None)

    for r in range(row, row + 3):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left   = med if c == col_start else no,
                right  = med if c == col_end   else no,
                top    = med if r == row        else no,
                bottom = med if r == row + 2    else no,
            )

    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    lc = ws.cell(row=row, column=col_start, value=label)
    lc.font = Font(bold=True, color="FFFFFF", size=9)
    lc.fill = bg_fill
    lc.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=row + 1, start_column=col_start,
                   end_row=row + 1, end_column=col_end)
    vc = ws.cell(row=row + 1, column=col_start, value=value)
    vc.font = Font(bold=True, size=20, color=bg_hex)
    vc.fill = light_fill
    vc.alignment = Alignment(horizontal="center", vertical="center")
    if number_format and isinstance(value, (int, float)):
        vc.number_format = number_format

    ws.merge_cells(start_row=row + 2, start_column=col_start,
                   end_row=row + 2, end_column=col_end)
    sc = ws.cell(row=row + 2, column=col_start, value=sublabel)
    sc.font = Font(size=9, italic=True, color="595959")
    sc.fill = light_fill
    sc.alignment = Alignment(horizontal="center", vertical="center")


def _mini_table_row(ws, row: int, label: str, values: list,
                    number_format: str, row_height: int = 20):
    """ミニテーブルの1データ行を書き込む"""
    ws.row_dimensions[row].height = row_height
    c = ws.cell(row=row, column=1, value=label)
    c.font = BOLD_FONT
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=i + 2, value=val)
        cell.number_format = number_format
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = BORDER


def _color_scale(ws, cell_range: str):
    ws.conditional_formatting.add(cell_range, ColorScaleRule(
        start_type="min",      start_color="FFC7CE",
        mid_type="percentile", mid_value=50, mid_color="E2E8F0",
        end_type="max",        end_color="C6EFCE",
    ))


def _write_ai_insight(ws, text: str, start_row: int = 21):
    """AI生成コメントを start_row 以降に1ブロック書き込む（凡例の下に配置）"""
    ws.row_dimensions[start_row - 1].height = 8  # スペーサー

    # ── 見出し ──
    ws.row_dimensions[start_row].height = 22
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=13)
    h = ws.cell(row=start_row, column=1, value="💡 AIインサイト（自動生成コメント）")
    h.font = Font(bold=True, size=11, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor="2E75B6")
    h.alignment = Alignment(horizontal="center", vertical="center")

    # ── 本文（折り返し表示）──
    body_start, body_end = start_row + 1, start_row + 5
    for r in range(body_start, body_end + 1):
        ws.row_dimensions[r].height = 18
    ws.merge_cells(start_row=body_start, start_column=1, end_row=body_end, end_column=13)
    c = ws.cell(row=body_start, column=1, value=text)
    c.font = Font(size=10, color="1F2937")
    c.fill = PatternFill("solid", fgColor="EBF3FB")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── 枠線（結合範囲の外周）──
    med, no = Side(style="medium"), Side(style=None)
    for r in range(body_start, body_end + 1):
        for col in range(1, 14):
            ws.cell(row=r, column=col).border = Border(
                left=med if col == 1 else no,
                right=med if col == 13 else no,
                top=med if r == body_start else no,
                bottom=med if r == body_end else no,
            )


# ── シート生成 ─────────────────────────────────────────────────────────────

def write_kpi_sheet(wb: Workbook, data: pd.DataFrame,
                    sales_pivot: pd.DataFrame, count_pivot: pd.DataFrame,
                    year: int):
    """先頭シート: KPIダッシュボード"""
    ws = wb.create_sheet("KPIダッシュボード", 0)
    ws.sheet_view.showGridLines = False

    # ── KPI計算 ──
    monthly_sales  = sales_pivot.sum(axis=1)
    monthly_counts = count_pivot.sum(axis=1)
    monthly_unit   = (monthly_sales / monthly_counts.replace(0, 1)).astype(int)
    total_annual   = int(monthly_sales.sum())

    best_m  = int(monthly_sales.idxmax())
    worst_m = int(monthly_sales.idxmin())
    avg_m   = int(monthly_sales.mean())

    store_annual    = sales_pivot.sum().sort_values(ascending=False)
    top_store       = store_annual.index[0]
    top_store_sales = int(store_annual.iloc[0])

    total_customers = int(count_pivot.values.sum())
    avg_unit_price  = int(total_annual / total_customers) if total_customers else 0

    dec = int(monthly_sales.loc[12])
    nov = int(monthly_sales.loc[11])
    dec_mom   = dec / nov if nov else 1.0
    # ↑=増加、▲=減少（日本の会計慣習に合わせ▲はマイナスを表す）
    mom_arrow = "↑" if dec_mom >= 1.0 else "▲"
    mom_str   = f"{mom_arrow} {abs(dec_mom - 1) * 100:.1f}%"
    mom_color = "375623" if dec_mom >= 1.0 else "843C0C"

    # ── 列幅（A-M = 1-13列）──
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 9
    ws.column_dimensions["A"].width = 16

    # ── Row 1: タイトル ──
    ws.row_dimensions[1].height = 38
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = f"{year}年  売上 KPI ダッシュボード"
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 8

    # ── Row 3-5: 年間総売上（メイン大カード）──
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 46
    ws.row_dimensions[5].height = 18
    _kpi_card(ws, 3, 1, 13, "年間総売上（全店舗合計）",
              total_annual,
              f"{year}年 1月〜12月  ※単位：{_yen_unit(total_annual)}",
              "1F4E79", _yen_fmt(total_annual))
    ws.cell(row=4, column=1).font = Font(bold=True, size=28, color="1F4E79")

    ws.row_dimensions[6].height = 8

    # ── Row 7-9: 最高月 / 最低月 / 月平均 ──
    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 34
    ws.row_dimensions[9].height = 18
    _best_s  = int(monthly_sales.loc[best_m])
    _worst_s = int(monthly_sales.loc[worst_m])
    _kpi_card(ws, 7, 1, 4, f"最高月売上（{_yen_unit(_best_s)}）",
              _best_s, MONTH_NAMES[best_m - 1], "375623", '#,##0,')
    _kpi_card(ws, 7, 5, 8, f"最低月売上（{_yen_unit(_worst_s)}）",
              _worst_s, MONTH_NAMES[worst_m - 1], "843C0C", '#,##0,')
    _kpi_card(ws, 7, 9, 13, f"月平均売上（{_yen_unit(avg_m)}）",
              avg_m, "12ヶ月平均", "404040", '#,##0,')

    ws.row_dimensions[10].height = 8

    # ── Row 11-13: No.1店舗 / 客単価 / 前月比 ──
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 34
    ws.row_dimensions[13].height = 18
    _kpi_card(ws, 11, 1, 4, "売上 No.1 店舗",
              top_store, f"年間 {_yen_str(top_store_sales)}", "1F4E79")
    _kpi_card(ws, 11, 5, 8, f"平均客単価（{_yen_unit(avg_unit_price)}）",
              avg_unit_price, "全店舗・全期間", "595959", _yen_fmt(avg_unit_price))
    _kpi_card(ws, 11, 9, 13, "12月 前月比",
              mom_str, f"12月 {_yen_str(dec)} ／ 11月 {_yen_str(nov)}", mom_color)

    ws.row_dimensions[14].height = 8

    # ── Row 15: ミニテーブルタイトル ──
    ws.row_dimensions[15].height = 22
    ws.merge_cells("A15:M15")
    c = ws["A15"]
    c.value = "月別全店推移"
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 16: ヘッダー ──
    ws.row_dimensions[16].height = 20
    h = ws.cell(row=16, column=1, value="項目")
    h.font = HEADER_FONT
    h.fill = PatternFill("solid", fgColor="2E75B6")
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border = BORDER
    for i, mn in enumerate(MONTH_NAMES):
        c = ws.cell(row=16, column=i + 2, value=mn)
        c.font = HEADER_FONT
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    # ── Row 17: 売上金額（千円）──
    _sales_vals = [int(monthly_sales.loc[m]) for m in range(1, 13)]
    _sales_fmt  = _yen_fmt(max(_sales_vals)).replace('"千円"', '').replace('"円"', '')
    _mini_table_row(ws, 17, "売上金額（千円）",
                    _sales_vals, _sales_fmt)
    for i, month in enumerate(range(1, 13)):
        col = i + 2
        ws.cell(row=17, column=col).font = BOLD_FONT

    # ── Row 18: 前月比 ──
    ws.row_dimensions[18].height = 20
    ws.cell(row=18, column=1, value="前月比").font = BOLD_FONT
    ws.cell(row=18, column=1).border = BORDER
    ws.cell(row=18, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for i, month in enumerate(range(1, 13)):
        col = i + 2
        if month == 1:
            c = ws.cell(row=18, column=col, value="-")
            c.alignment = Alignment(horizontal="center")
        else:
            prev = int(monthly_sales.loc[month - 1])
            curr = int(monthly_sales.loc[month])
            mom  = curr / prev - 1 if prev else 0.0
            c = ws.cell(row=18, column=col, value=round(mom, 4))
            c.number_format = '+0.0%;-0.0%;"-"'
            c.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=18, column=col).border = BORDER

    _apply_mom_cf(ws, f"C18:{get_column_letter(13)}18")

    # ── Row 19: 凡例（前月比の色分けルール説明、表の直下に配置）──
    ws.row_dimensions[19].height = 22

    legend_label = ws.cell(row=19, column=1, value="前月比 凡例")
    legend_label.font = Font(bold=True, size=10, color="1F4E79")
    legend_label.fill = PatternFill("solid", fgColor="DEEAF1")
    legend_label.alignment = Alignment(horizontal="center", vertical="center")
    legend_label.border = BORDER

    # 通常セル用の静的塗り（CF は bgColor、通常セルは solid fgColor が必要）
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="E2E8F0")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")

    # 緑: +5%以上（増加）
    ws.merge_cells(start_row=19, start_column=3, end_row=19, end_column=4)
    g = ws.cell(row=19, column=3, value="  ↑  +5%以上（増加）")
    g.fill = green_fill
    g.font = Font(color="276221", size=9, bold=True)
    g.alignment = Alignment(horizontal="center", vertical="center")
    g.border = BORDER

    # 黄: ±5%以内（横ばい）
    ws.merge_cells(start_row=19, start_column=6, end_row=19, end_column=8)
    y = ws.cell(row=19, column=6, value="  ±5%以内（横ばい）")
    y.fill = yellow_fill
    y.font = Font(color="475569", size=9, bold=True)
    y.alignment = Alignment(horizontal="center", vertical="center")
    y.border = BORDER

    # 赤: ▲ -5%以下（減少）
    ws.merge_cells(start_row=19, start_column=10, end_row=19, end_column=12)
    r = ws.cell(row=19, column=10, value="  ▲  -5%以下（減少）")
    r.fill = red_fill
    r.font = Font(color="9C0006", size=9, bold=True)
    r.alignment = Alignment(horizontal="center", vertical="center")
    r.border = BORDER

    # 凡例行の全セルに枠線を付与（結合末尾セルの右枠含む）
    for col in range(1, 14):
        ws.cell(row=19, column=col).border = BORDER

    # ── AIインサイト（任意・環境変数 ANTHROPIC_API_KEY がある場合のみ）──
    kpi_summary = {
        "year": year,
        "total_annual": total_annual,
        "monthly_sales": _sales_vals,
        "best_month": best_m,   "best_sales": _best_s,
        "worst_month": worst_m, "worst_sales": _worst_s,
        "avg_month": avg_m,
        "store_ranking": [(store, int(sales)) for store, sales in store_annual.items()],
        "avg_unit_price": avg_unit_price,
        "total_customers": total_customers,
        "last_month_mom": dec_mom - 1.0,
    }
    commentary = generate_kpi_commentary(kpi_summary)
    if commentary:
        _write_ai_insight(ws, commentary)


def write_summary_sheet(wb: Workbook, sales_pivot: pd.DataFrame,
                        count_pivot: pd.DataFrame, year: int):
    ws = wb.create_sheet("月別集計サマリー")
    stores   = list(sales_pivot.columns)
    n_stores = len(stores)

    ws.merge_cells(f"A1:{get_column_letter(n_stores * 2 + 3)}1")
    title_cell = ws["A1"]
    title_cell.value = f"{year}年 月別売上集計レポート"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value="月")
    col = 2
    for store in stores:
        ws.cell(row=2, column=col,     value=f"{store}\n売上金額（千円）")
        ws.cell(row=2, column=col + 1, value=f"{store}\n客数（人）")
        col += 2
    ws.cell(row=2, column=col,     value="全店合計\n売上金額（千円）")
    ws.cell(row=2, column=col + 1, value="全店合計\n客数（人）")

    style_header_row(ws, 2, 1, n_stores * 2 + 3)
    ws.row_dimensions[2].height = 30
    ws.column_dimensions["A"].width = 9

    for i, month in enumerate(range(1, 13)):
        row    = 3 + i
        is_alt = i % 2 == 1
        ws.cell(row=row, column=1, value=MONTH_NAMES[month - 1])
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = BORDER

        col = 2
        for store in stores:
            s_cell = ws.cell(row=row, column=col,
                             value=int(sales_pivot.loc[month, store]))
            c_cell = ws.cell(row=row, column=col + 1,
                             value=int(count_pivot.loc[month, store]))
            style_data_cell(s_cell, is_alt, '#,##0,')
            style_data_cell(c_cell, is_alt, '#,##0')
            col += 2

        total_sales = int(sales_pivot.loc[month].sum())
        total_count = int(count_pivot.loc[month].sum())
        ts_cell = ws.cell(row=row, column=col,     value=total_sales)
        tc_cell = ws.cell(row=row, column=col + 1, value=total_count)
        ts_cell.font = BOLD_FONT
        tc_cell.font = BOLD_FONT
        style_data_cell(ts_cell, is_alt, '#,##0,')
        style_data_cell(tc_cell, is_alt, '#,##0')

    total_row = 15
    ws.cell(row=total_row, column=1, value="年間合計")
    ws.cell(row=total_row, column=1).font  = BOLD_FONT
    ws.cell(row=total_row, column=1).fill  = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = BORDER
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

    col = 2
    for store in stores:
        s = ws.cell(row=total_row, column=col,
                    value=int(sales_pivot[store].sum()))
        c = ws.cell(row=total_row, column=col + 1,
                    value=int(count_pivot[store].sum()))
        s.font = BOLD_FONT; s.fill = TOTAL_FILL; s.border = BORDER
        s.number_format = '#,##0,';   s.alignment = Alignment(horizontal="right")
        c.font = BOLD_FONT; c.fill = TOTAL_FILL; c.border = BORDER
        c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
        col += 2

    s = ws.cell(row=total_row, column=col,     value=int(sales_pivot.values.sum()))
    c = ws.cell(row=total_row, column=col + 1, value=int(count_pivot.values.sum()))
    for cell, fmt in [(s, '#,##0,'), (c, '#,##0')]:
        cell.font = Font(bold=True, color="1F4E79")
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")

    col_widths = {2: 22.64, 3: 15.82, 4: 22.64, 5: 15.82,
                  6: 22.64, 7: 15.82, 8: 24.91, 9: 18.09}
    for col_idx in range(2, n_stores * 2 + 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 16)

    cats = Reference(ws, min_col=1, min_row=3, max_row=14)
    store_colors = ["4472C4", "ED7D31", "70AD47", "FFC000", "5B9BD5"]

    def _setup_axis(ch):
        ch.x_axis.delete = False
        ch.x_axis.tickLblPos = "nextTo"
        ch.y_axis.delete = False
        ch.y_axis.tickLblPos = "nextTo"
        ch.y_axis.numFmt = '#,##0,'
        ch.y_axis.sourceLinked = False
        ch.plot_area.layout = Layout(
            manualLayout=ManualLayout(layoutTarget="inner", x=0.12, y=0.06, w=0.82, h=0.82)
        )

    # ── グラフ1: 全店合計売上（棒グラフ）────────────────────────────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "月別 全店合計売上"
    chart1.y_axis.title = "売上金額（千円）"
    chart1.x_axis.title = None
    chart1.style = 2
    chart1.width = 13.6
    chart1.height = 8
    chart1.legend = None

    data_ref = Reference(ws, min_col=n_stores * 2 + 2, max_col=n_stores * 2 + 2,
                         min_row=3, max_row=14)
    chart1.add_data(data_ref, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.solidFill = "4472C4"
    chart1.series[0].graphicalProperties.line.solidFill = "4472C4"

    _setup_axis(chart1)
    chart1.plot_area.spPr = GraphicalProperties(
        ln=LineProperties(solidFill='A6A6A6', w=9525)
    )
    ws.add_chart(chart1, f"A{total_row + 2}")

    # ── グラフ2: 集合棒グラフ（店舗別売上比較）──────────────────────────
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "月別 店舗別売上比較"
    chart2.y_axis.title = "売上金額（千円）"
    chart2.x_axis.title = None
    chart2.style = 2
    chart2.width = 18.86
    chart2.height = 8

    for i, store in enumerate(stores):
        col_idx = 2 + i * 2
        data_ref = Reference(ws, min_col=col_idx, max_col=col_idx, min_row=3, max_row=14)
        chart2.add_data(data_ref, titles_from_data=False)
        chart2.series[i].title = SeriesLabel(v=store)
        color = store_colors[i % len(store_colors)]
        chart2.series[i].graphicalProperties.solidFill = color
        chart2.series[i].graphicalProperties.line.solidFill = color

    chart2.set_categories(cats)
    _setup_axis(chart2)
    chart2.plot_area.spPr = GraphicalProperties(
        ln=LineProperties(solidFill='A6A6A6', w=9525)
    )
    ws.add_chart(chart2, f"E{total_row + 2}")


def write_store_sheet(wb: Workbook, store_name: str,
                      store_data: pd.DataFrame, year: int, sheet_title: str = None):
    safe_name = sheet_title if sheet_title else store_name.replace("/", "_")
    ws = wb.create_sheet(safe_name)

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"{year}年 {store_name} 月別売上詳細"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["月", "売上金額（千円）", "客数（人）", "客単価（円）", "前月比", "累計売上（千円）"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[2].height = 22

    monthly = store_data.groupby("月").agg(
        売上金額=("売上金額", "sum"),
        客数=("客数", "sum"),
    ).reindex(range(1, 13), fill_value=0)
    monthly["客単価"] = (monthly["売上金額"] / monthly["客数"].replace(0, 1)).astype(int)
    monthly["前月比"] = monthly["売上金額"] / monthly["売上金額"].shift(1) - 1
    monthly["累計売上"] = monthly["売上金額"].cumsum()

    for i, month in enumerate(range(1, 13)):
        row    = 3 + i
        is_alt = i % 2 == 1
        r = monthly.loc[month]

        ws.cell(row=row, column=1,
                value=MONTH_NAMES[month - 1]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = BORDER

        s_cell   = ws.cell(row=row, column=2, value=int(r["売上金額"]))
        c_cell   = ws.cell(row=row, column=3, value=int(r["客数"]))
        p_cell   = ws.cell(row=row, column=4, value=int(r["客単価"]))
        mom_cell = ws.cell(row=row, column=5,
                           value=None if month == 1 else round(float(r["前月比"]), 4))
        cum_cell = ws.cell(row=row, column=6, value=int(r["累計売上"]))

        style_data_cell(s_cell,   is_alt, '#,##0,')
        style_data_cell(c_cell,   is_alt, '#,##0')
        style_data_cell(p_cell,   is_alt, '#,##0')
        style_data_cell(cum_cell, is_alt, '#,##0,')
        mom_cell.number_format = '+0.0%;-0.0%;"-"'
        mom_cell.alignment = Alignment(horizontal="right")
        mom_cell.border = BORDER
        if is_alt:
            mom_cell.fill = ALT_ROW_FILL

    # 前月比列（E4:E14）に条件付き書式
    _apply_mom_cf(ws, "E4:E14")

    total_row = 15
    ws.cell(row=total_row, column=1, value="年間合計").font = BOLD_FONT
    ws.cell(row=total_row, column=1).fill   = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = BORDER
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")

    for col, val, fmt in [
        (2, int(monthly["売上金額"].sum()), '#,##0,'),
        (3, int(monthly["客数"].sum()),     '#,##0'),
        (4, int(monthly["売上金額"].sum() / monthly["客数"].sum()), '#,##0'),
        (6, int(monthly["売上金額"].sum()), '#,##0,'),
    ]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")

    widths = [9, 18, 12, 14, 12, 20]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    chart = LineChart()
    chart.title = f"{store_name} 月別売上推移"
    chart.y_axis.title = "売上金額（千円）"
    chart.y_axis.numFmt = '#,##0,'
    chart.y_axis.sourceLinked = False
    chart.x_axis.title = None
    chart.style = 10
    chart.width = sum(w * 7 + 5 for w in widths) / 96 * 2.54  # align to columns A-F
    chart.height = 10
    chart.legend = None
    chart.plot_area.spPr = GraphicalProperties(
        ln=LineProperties(solidFill='A6A6A6', w=9525)
    )

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=14)
    cats = Reference(ws, min_col=1, min_row=3, max_row=14)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{total_row + 2}")


# ── GUI用公開関数 ────────────────────────────────────────────────────────────

def build_base_workbook(db_path: Path, year: int, output_path: Path) -> None:
    """KPIダッシュボードと月別集計サマリーのみを生成して保存する"""
    data = load_from_db(db_path, year)
    sales_pivot, count_pivot = build_monthly_pivot(data)
    wb = Workbook()
    wb.remove(wb.active)
    write_kpi_sheet(wb, data, sales_pivot, count_pivot, year)
    write_summary_sheet(wb, sales_pivot, count_pivot, year)
    wb.save(output_path)


def _reapply_plot_area_borders(wb) -> None:
    """load_workbook後にplot_area.spPrが失われるopenpyxlのバグを回避する"""
    border = GraphicalProperties(ln=LineProperties(solidFill='A6A6A6', w=9525))
    for ws in wb.worksheets:
        for chart in ws._charts:
            chart.plot_area.spPr = border


def update_store_sheet(db_path: Path, year: int, output_path: Path, store_name: str) -> None:
    """既存Excelの「店舗詳細」シートのみを差し替えて上書き保存する"""
    from openpyxl import load_workbook as _load_workbook
    data = load_from_db(db_path, year)
    store_data = data[data["店舗名"] == store_name].copy()
    wb = _load_workbook(output_path)
    if "店舗詳細" in wb.sheetnames:
        del wb["店舗詳細"]
    write_store_sheet(wb, store_name, store_data, year, sheet_title="店舗詳細")
    _reapply_plot_area_borders(wb)
    wb.save(output_path)


# ── main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="月別売上集計レポート自動生成")
    parser.add_argument("--input",  default="data/sample",                help="CSVフォルダパス")
    parser.add_argument("--output", default="output/monthly_report.xlsx", help="出力Excelパス")
    parser.add_argument("--year",   type=int, default=2024,               help="対象年")
    args = parser.parse_args()

    input_dir   = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(exist_ok=True)

    print(f"CSVを読み込み中: {input_dir}")
    data = load_csv_files(input_dir, args.year)
    stores = sorted(data["店舗名"].unique())
    print(f"店舗数: {len(stores)}  行数: {len(data):,}")

    sales_pivot, count_pivot = build_monthly_pivot(data)

    wb = Workbook()
    wb.remove(wb.active)

    print("KPIダッシュボード作成中...")
    write_kpi_sheet(wb, data, sales_pivot, count_pivot, args.year)

    print("月別集計サマリーシート作成中...")
    write_summary_sheet(wb, sales_pivot, count_pivot, args.year)

    for store in stores:
        print(f"  {store} シート作成中...")
        write_store_sheet(wb, store, data[data["店舗名"] == store], args.year)

    wb.save(output_path)
    print(f"\n完了: {output_path}")
    print(f"シート: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
